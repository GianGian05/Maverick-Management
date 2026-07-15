from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import date, datetime
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.properties import CalcProperties
from PIL import Image, ImageOps

from prepare_rfp_template import all_paragraphs, replace_in_paragraph


ROOT = Path(__file__).resolve().parent
SCHEMA_PATH = ROOT / "roof_schema.json"
TEMPLATE_DIR = ROOT / "templates"
READY_FILL = PatternFill("solid", fgColor="E2F0D9")
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")


def read_schema() -> dict:
    return json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_project_intake(path: Path) -> tuple[dict, dict]:
    schema = read_schema()
    workbook = load_workbook(path, data_only=True)
    if "Project Input" not in workbook.sheetnames:
        raise ValueError("The intake workbook is missing the 'Project Input' sheet.")
    sheet = workbook["Project Input"]
    data = {}
    for offset, field in enumerate(schema["project_fields"], start=5):
        key, label, default, required, _ = field
        value = sheet.cell(offset, 2).value
        data[key] = default if value is None and not required else value
        if sheet.cell(offset, 1).value != label:
            raise ValueError(f"Project Input row {offset} was changed. Expected field '{label}'.")
    data["full_address"] = f"{data['address_line_1']}, {data['city']}, {data['state']} {data['zip']}"
    data["city_state_zip"] = f"{data['city']}, {data['state']} {data['zip']}"
    return data, schema


def validate_project(data: dict, schema: dict) -> list[str]:
    warnings = []
    required = {key: label for key, label, _, is_required, _ in schema["project_fields"] if is_required}
    for key, label in required.items():
        if data.get(key) in (None, ""):
            warnings.append(f"Missing required field: {label}")
    state_names = {
        "MN": "Minnesota", "CT": "Connecticut", "NY": "New York", "GA": "Georgia",
        "TX": "Texas", "CA": "California", "FL": "Florida", "IL": "Illinois",
        "TN": "Tennessee", "VA": "Virginia", "PA": "Pennsylvania", "NJ": "New Jersey",
    }
    expected_license = state_names.get(str(data.get("state", "")).upper())
    if expected_license and str(data.get("license_state", "")).strip().lower() != expected_license.lower():
        warnings.append(f"License state '{data.get('license_state')}' does not match project state '{expected_license}'.")
    issue = parse_datetime(data.get("rfp_issue_date"))
    due = parse_datetime(data.get("bid_due"))
    if issue and due and due <= issue:
        warnings.append("Bid due date must be after the RFP issue date.")
    if data.get("area_sf") is not None:
        try:
            if float(data["area_sf"]) <= 0:
                warnings.append("Roof area must be greater than zero.")
        except (TypeError, ValueError):
            warnings.append("Roof area must be numeric.")
    return warnings


def parse_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if not value:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %I:%M %p", "%Y-%m-%d", "%m/%d/%Y %I:%M %p", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_date(value, long: bool = False) -> str:
    parsed = parse_datetime(value)
    if not parsed:
        return str(value or "")
    return parsed.strftime("%B %-d, %Y" if long and sys.platform != "win32" else "%B %#d, %Y" if long else "%m/%d/%Y")


def format_bid_due(value) -> str:
    parsed = parse_datetime(value)
    if not parsed:
        return str(value or "")
    day = parsed.strftime("%B %d, %Y").replace(" 0", " ")
    clock = parsed.strftime("%-I:%M %p" if sys.platform != "win32" else "%#I:%M %p")
    return f"{day}, {clock}"


def format_currency(value) -> str:
    try:
        return f"${float(value):,.2f}"
    except (TypeError, ValueError):
        return str(value or "")


def replace_cover_image(docx_path: Path, image_path: Path) -> None:
    with Image.open(image_path) as source:
        source = ImageOps.exif_transpose(source).convert("RGB")
        fitted = ImageOps.fit(source, (1250, 720), method=Image.Resampling.LANCZOS)
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as handle:
            temp_image = Path(handle.name)
        fitted.save(temp_image, "PNG", optimize=True)
    temp_docx = docx_path.with_suffix(".tmp.docx")
    with zipfile.ZipFile(docx_path, "r") as source_zip, zipfile.ZipFile(temp_docx, "w", zipfile.ZIP_DEFLATED) as output_zip:
        for item in source_zip.infolist():
            payload = temp_image.read_bytes() if item.filename == "word/media/image1.png" else source_zip.read(item.filename)
            output_zip.writestr(item, payload)
    temp_docx.replace(docx_path)
    temp_image.unlink(missing_ok=True)


def generate_rfp(intake: Path, output_dir: Path, make_pdf: bool) -> Path:
    data, schema = load_project_intake(intake)
    warnings = validate_project(data, schema)
    if warnings:
        raise ValueError("Project validation failed:\n- " + "\n- ".join(warnings))

    manifest_path = TEMPLATE_DIR / "RFP_Template_Manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    template = TEMPLATE_DIR / manifest["template_file"]
    if file_sha256(template) != manifest["sha256"]:
        raise ValueError("The locked RFP template has been modified. Restore the approved template before generating an RFP.")

    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", str(data["project_name"])).strip("_")
    output = output_dir / f"RFP_{safe_name}.docx"
    shutil.copy2(template, output)
    doc = Document(output)
    replacements = {
        "{{PROJECT_NAME}}": str(data["project_name"]),
        "{{ROOF_PROJECT_TYPE}}": str(data["roof_project_type"]),
        "{{OWNER_ENTITY}}": str(data["owner_entity"]),
        "{{OWNER_ENTITY_LEGAL}}": str(data["owner_entity_legal"]),
        "{{ADDRESS_LINE_1}}": str(data["address_line_1"]),
        "{{FULL_ADDRESS}}": str(data["full_address"]),
        "{{CITY_STATE_ZIP}}": str(data["city_state_zip"]),
        "{{AREA_SF}}": f"{float(data['area_sf']):,.0f}",
        "{{RFP_ISSUE_DATE_LONG}}": format_date(data["rfp_issue_date"], long=True),
        "{{RFP_ISSUE_DATE_SHORT}}": format_date(data["rfp_issue_date"], long=False),
        "{{JOB_WALK}}": str(data["job_walk"]),
        "{{BID_DUE}}": format_bid_due(data["bid_due"]),
        "{{LICENSE_STATE}}": str(data["license_state"]),
        "{{PRIMARY_CONTACT_NAME}}": str(data["primary_contact_name"]),
        "{{PRIMARY_CONTACT_EMAIL}}": str(data["primary_contact_email"]),
        "{{COPY_EMAILS}}": str(data.get("copy_emails") or "").replace(";", ","),
        "{{OPTION_A_NAME}}": str(data["option_a_name"]),
        "{{OPTION_B_NAME}}": str(data["option_b_name"]),
        "{{LIABILITY_LIMIT}}": format_currency(data["liability_limit"]),
    }
    for token, value in replacements.items():
        for paragraph in all_paragraphs(doc):
            replace_in_paragraph(paragraph, token, value)
    unresolved = []
    for paragraph in all_paragraphs(doc):
        unresolved.extend(re.findall(r"\{\{[^}]+\}\}", paragraph.text))
    if unresolved:
        raise ValueError(f"Unresolved RFP placeholders: {sorted(set(unresolved))}")
    doc.save(output)

    cover = str(data.get("cover_image_path") or "").strip()
    if cover:
        cover_path = Path(cover)
        if not cover_path.is_absolute():
            cover_path = (intake.parent / cover_path).resolve()
        if not cover_path.exists():
            raise FileNotFoundError(f"Cover image not found: {cover_path}")
        replace_cover_image(output, cover_path)

    if make_pdf:
        try:
            from docx2pdf import convert
            convert(str(output), str(output.with_suffix(".pdf")))
        except Exception as exc:
            print(f"PDF conversion skipped: {exc}. Open the DOCX in Word and choose Save as PDF.", file=sys.stderr)
    return output


def generate_bid_form(intake: Path, output_dir: Path) -> Path:
    data, schema = load_project_intake(intake)
    warnings = validate_project(data, schema)
    if warnings:
        raise ValueError("Project validation failed:\n- " + "\n- ".join(warnings))
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / "Contractor_Roof_Bid_Form.xlsx"
    shutil.copy2(TEMPLATE_DIR / "Contractor_Roof_Bid_Form.xlsx", output)
    wb = load_workbook(output)
    ws = wb["Contractor Bid Form"]
    ws["B5"] = data["project_name"]
    ws["B6"] = data["full_address"]
    ws["B7"] = data["owner_entity"]
    ws["B8"] = format_bid_due(data["bid_due"])
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    wb.save(output)
    return output


def extract_bid(path: Path) -> tuple[dict, list[str], str]:
    wb = load_workbook(path, data_only=True)
    if "Automation Map" not in wb.sheetnames:
        raise ValueError(f"{path.name}: missing Automation Map sheet")
    map_sheet = wb["Automation Map"]
    version = str(map_sheet["B2"].value or "")
    result = {}
    warnings = []
    row = 4
    while map_sheet.cell(row, 1).value:
        key = str(map_sheet.cell(row, 1).value)
        sheet_name = str(map_sheet.cell(row, 2).value)
        cell = str(map_sheet.cell(row, 3).value)
        if sheet_name not in wb.sheetnames:
            warnings.append(f"Missing sheet: {sheet_name}")
            result[key] = None
        else:
            result[key] = wb[sheet_name][cell].value
        row += 1
    if not result.get("company"):
        warnings.append("Bidder company is blank")
    if result.get("option_a_base_bid") in (None, "") and result.get("option_b_base_bid") in (None, ""):
        warnings.append("Both base bid options are blank")
    blank_scope = sum(1 for ref in range(1, 47) if result.get(f"scope_{ref}_include") in (None, ""))
    if blank_scope:
        warnings.append(f"{blank_scope} scope inclusion responses are blank")
    exclusions = sum(1 for ref in range(1, 47) if str(result.get(f"scope_{ref}_include") or "").strip().upper() in ("N", "NO"))
    if exclusions:
        warnings.append(f"{exclusions} scope item(s) are excluded")
    return result, warnings, version


def find_label_row(sheet, label: str) -> int:
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row, 1).value or "").strip() == label:
            return row
    raise KeyError(f"Leveling template label not found: {label}")


def level_bids(intake: Path, bids: list[Path], output: Path) -> Path:
    if not 1 <= len(bids) <= 4:
        raise ValueError("Provide between 1 and 4 contractor bid workbooks.")
    data, schema = load_project_intake(intake)
    warnings = validate_project(data, schema)
    if warnings:
        raise ValueError("Project validation failed:\n- " + "\n- ".join(warnings))
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE_DIR / "Roof_Bid_Leveling_Template.xlsx", output)
    wb = load_workbook(output)
    summary = wb["Bid Leveling - Summary"]
    comparison = wb["Scope Comparison"]
    audit = wb["Import Audit"]
    summary["B3"] = data["owner_entity"]
    summary["B4"] = data["project_name"]
    summary["B5"] = data["full_address"]
    summary["B6"] = format_bid_due(data["bid_due"])
    summary["B7"] = float(data["area_sf"])
    summary["B8"] = schema["template_version"]

    field_groups = [schema["bidder_information"], schema["company_profile"], schema["pricing_fields"], schema["alternates"], schema["unit_prices"]]
    fields = [item for group in field_groups for item in group]
    for idx, bid_path in enumerate(bids):
        bid, bid_warnings, version = extract_bid(bid_path)
        col = 3 + idx
        company = str(bid.get("company") or f"Bidder {idx + 1}")
        summary.cell(10, col).value = company
        for key, label in fields:
            row = find_label_row(summary, label)
            summary.cell(row, col).value = bid.get(key)
        include_col = 3 + idx * 2
        remark_col = include_col + 1
        comparison.cell(3, include_col).value = f"{company} Include"
        comparison.cell(3, remark_col).value = f"{company} Remark"
        for scope_idx, (ref, _) in enumerate(schema["scope_items"], start=4):
            comparison.cell(scope_idx, include_col).value = bid.get(f"scope_{ref}_include")
            comparison.cell(scope_idx, remark_col).value = bid.get(f"scope_{ref}_remark")
            if bid.get(f"scope_{ref}_include") in (None, "", "N", "No"):
                comparison.cell(scope_idx, include_col).fill = REVIEW_FILL
        audit_row = 4 + idx
        audit.cell(audit_row, 1).value = bid_path.name
        audit.cell(audit_row, 2).value = company
        audit.cell(audit_row, 3).value = version
        audit.cell(audit_row, 4).value = sum(1 for value in bid.values() if value not in (None, ""))
        audit.cell(audit_row, 5).value = "; ".join(bid_warnings)
        status = "READY" if not bid_warnings and version == schema["template_version"] else "REVIEW"
        audit.cell(audit_row, 6).value = status
        audit.cell(audit_row, 6).fill = READY_FILL if status == "READY" else REVIEW_FILL
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    wb.save(output)
    return output


def command_prepare(args) -> None:
    intake = Path(args.intake).resolve()
    output_dir = Path(args.output_dir).resolve()
    print(f"Created {generate_rfp(intake, output_dir, args.pdf)}")
    print(f"Created {generate_bid_form(intake, output_dir)}")


def command_level(args) -> None:
    output = Path(args.output).resolve()
    bids = [Path(item).resolve() for item in args.bids]
    print(f"Created {level_bids(Path(args.intake).resolve(), bids, output)}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Maverick roof RFP and bid-leveling automation")
    subparsers = parser.add_subparsers(dest="command", required=True)
    prepare = subparsers.add_parser("prepare", help="Generate the RFP and contractor bid form")
    prepare.add_argument("--intake", required=True, help="Roof_Project_Intake.xlsx")
    prepare.add_argument("--output-dir", required=True)
    prepare.add_argument("--pdf", action="store_true", help="Attempt PDF conversion using Microsoft Word")
    prepare.set_defaults(func=command_prepare)
    level = subparsers.add_parser("level", help="Import contractor Excel bids into the leveling workbook")
    level.add_argument("--intake", required=True)
    level.add_argument("--bids", nargs="+", required=True, help="One to four completed contractor bid workbooks")
    level.add_argument("--output", required=True)
    level.set_defaults(func=command_level)
    return parser


if __name__ == "__main__":
    arguments = build_parser().parse_args()
    arguments.func(arguments)
